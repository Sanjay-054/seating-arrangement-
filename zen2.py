def missing(missing,st_list):
    for i in missing:
         for j in st_list:
            if (j  == i):
                st_list.remove(i)
    return st_list

def run(bcom_list, bsc_list, bca_list, bba_list, col, row, subject_code_bca, subject_code_bsc, subject_code_bcom, subject_code_bba):
    data = {
        "row": row,
        "col": col,
        "BCA": {"students": bca_list, "subject_code": subject_code_bca},
        "BSc": {"students": bsc_list, "subject_code": subject_code_bsc},
        "BCom": {"students": bcom_list, "subject_code": subject_code_bcom},
        "BBA": {"students": bba_list, "subject_code": subject_code_bba},
    }
    return data
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook  # For writing to Excel using openpyxl
import random
from openpyxl.styles import Font,Alignment,PatternFill
from openpyxl.utils import get_column_letter

# Function to write matrix-style data to an Excel file
def write(result, room_no,subject_code):
    # Ensure the result is structured correctly
    print("Result Shape:", len(result), "x", len(result[0]) if result else 0)
    
    df = pd.DataFrame(result)

    # Ensure the file path is formatted correctly
    file_path = f'static/excel/{room_no}_{subject_code}.xlsx'
    
    # Create an Excel writer object
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')

    # Save the DataFrame to an Excel sheet
    df.to_excel(writer, sheet_name=str(room_no), header=False, index=False)

    # Close the writer (important for saving changes)
    writer.close()

    print(f"Excel file saved: {file_path}")

# Function to write detailed data (with subject code and other details) to Excel
def write_to_excel(data, room_no,session,date,shift):
    workbook = Workbook()
    sheet = workbook.active

    print("Data received:", data)

    # **Insert College Name in G1:L1**
    sheet.merge_cells("G1:L1")  
    sheet["G1"] = "Patrician College of Arts and Science"
    sheet["G1"].font = Font(bold=True, size=14)
    sheet["G1"].alignment = Alignment(horizontal="center", vertical="center")  # ✅ **Fixed Centering**

    # **Insert Room, Session, Date, and Shift**
    sheet.append(["Room Number", "Session", "Date", "Shift", "Rows", "Columns"])
    sheet.append([room_no, session, date, shift, data.get("row"), data.get("col")])
    sheet.append([])  

    for dept, details in data.items():
        if dept in ["row", "col"]:
            continue
        
        subject_code = details["subject_code"]
        students = details["students"]

        # **Department Header**
        sheet.append([f"{dept.upper()} Students - Subject Code: {subject_code}"])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
        sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(horizontal="center", vertical="center")  # ✅ **Centered Header**

        if students:
            # **Header for Student Roll Numbers**
            sheet.append(["Student Roll Numbers"])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
            sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

            # **Distribute Student Roll Numbers Across Columns**
            num_columns = 10  # Adjust as needed
            for i in range(0, len(students), num_columns):
                sheet.append(students[i:i+num_columns])  # Add students in rows of 10
                last_row = sheet.max_row  

                # ✅ **Ensure Center Alignment for Roll Numbers**
                for col_idx in range(1, len(students[i:i+num_columns]) + 1):
                    sheet.cell(row=last_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

            sheet.append([])  # Blank row for spacing
        else:
            sheet.append(["No students available"])
            sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # **Auto-adjust column width (Fixed Empty Sequence Issue)**
    for col in sheet.columns:
        col_letter = get_column_letter(col[0].column)
        col_values = [str(cell.value) for cell in col if cell.value]
        max_length = max((len(value) for value in col_values), default=10)  # ✅ **Avoids Empty Sequence Error**
        sheet.column_dimensions[col_letter].width = max_length + 2  

    # **Save the file**
    save_dir = "static/excel"
    os.makedirs(save_dir, exist_ok=True)

    filename = f'seating_{room_no}_{session}_{date}_{shift}.xlsx'.replace(" ", "_")
    filepath = os.path.join(save_dir, filename)
    workbook.save(filepath)

    print(f"File saved: {filepath}")
    return filename
# Function to read from an Excel file
def read(room_no,session):
    # Set display format for numbers
    pd.options.display.float_format = '{:,.0f}'.format
    filename=f'static/excel/seating_{room_no}_{session}.xlsx'
    filename2=f'static/excel/{room_no}.xlsx',
    # Read the Excel file using openpyxl as the engine
    temp = pd.read_excel( filename,header=None, index_col=False, engine='openpyxl')

    # Replace NaN values with "Blank"
    temp = temp.fillna('.')

    # Optionally, replace zeros with "Blank"
    temp = temp.replace(0, '.')

    # Convert all values to strings
    temp = temp.astype(str)

    return temp

